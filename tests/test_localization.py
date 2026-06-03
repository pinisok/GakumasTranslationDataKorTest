"""Tests for Localization pipeline."""

import os
import json

import pytest

from tests.fixtures.create_fixtures import create_localization_xlsx
from scripts.localization import (
    XlsxToJson as LocXlsxToJson,
    UpdateOriginalToDrive,
    ConvertDriveToOutput,
)
from scripts import localization_release
from scripts.localization_release import (
    LocalizationDiff,
    ReleaseInfo,
    OBSOLETE_MARKER,
    _parse_release_payload,
    diff_release_against_xlsx,
    apply_diff_to_xlsx,
    append_release_notes,
    load_last_release_tag,
    save_release_tag,
)


class TestLocalizationXlsxToJson:
    """Localization XlsxToJson with synthetic fixtures."""

    def test_basic_conversion(self, tmp_path):
        xlsx_path = str(tmp_path / "localization.xlsx")
        create_localization_xlsx(xlsx_path, [
            {0: "ref1", "ID": "ui.button.ok", "번역": "확인"},
            {0: "ref2", "ID": "ui.button.cancel", "번역": "취소"},
        ])
        output_path = str(tmp_path / "localization.json")

        LocXlsxToJson(xlsx_path, output_path)

        with open(output_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        assert data["ui.button.ok"] == "확인"
        assert data["ui.button.cancel"] == "취소"

    def test_output_keys_are_id_column(self, tmp_path):
        xlsx_path = str(tmp_path / "localization.xlsx")
        create_localization_xlsx(xlsx_path, [
            {0: "original_ref", "ID": "my.key", "번역": "값"},
        ])
        output_path = str(tmp_path / "localization.json")

        LocXlsxToJson(xlsx_path, output_path)

        with open(output_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        assert "my.key" in data
        assert "original_ref" not in data

    def test_leading_apostrophe_stripped(self, tmp_path):
        xlsx_path = str(tmp_path / "localization.xlsx")
        create_localization_xlsx(xlsx_path, [
            {0: "ref", "ID": "key", "번역": "\'quoted"},
        ])
        output_path = str(tmp_path / "localization.json")

        LocXlsxToJson(xlsx_path, output_path)

        with open(output_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        assert data["key"] == "quoted"

    def test_deserialize_applied(self, tmp_path):
        xlsx_path = str(tmp_path / "localization.xlsx")
        create_localization_xlsx(xlsx_path, [
            {0: "ref", "ID": "key", "번역": "line1\\tline2"},
        ])
        output_path = str(tmp_path / "localization.json")

        LocXlsxToJson(xlsx_path, output_path)

        with open(output_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        assert data["key"] == "line1\tline2"


class TestLocalizationUpdate:
    def test_update_skips_when_release_unavailable(self, monkeypatch):
        monkeypatch.setattr(
            "scripts.localization.localization_release.fetch_latest_release",
            lambda *args, **kwargs: None,
        )
        file_list, warnings = UpdateOriginalToDrive()
        assert file_list == []
        assert warnings == {}

    def test_incremental_skips(self):
        errors, successes = ConvertDriveToOutput(drive_file_paths=[])
        assert errors == []
        assert successes == []


class TestLocalizationNumericTranslation:
    """Numeric value in 번역 column should be skipped."""

    def test_numeric_translation_skipped(self, tmp_path):
        xlsx_path = str(tmp_path / "localization.xlsx")
        create_localization_xlsx(xlsx_path, [
            {0: "ref1", "ID": "key.valid", "번역": "정상번역"},
            {0: "ref2", "ID": "key.numeric", "번역": 12345},
        ])
        output_path = str(tmp_path / "localization.json")

        LocXlsxToJson(xlsx_path, output_path)

        with open(output_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        assert "key.valid" in data
        assert "key.numeric" not in data


# ============================================================
# P3: load_cache_date multi-line file (P3-20)
# ============================================================


# ============================================================
# Release-driven update flow (pinisok/gaku-patcher localization.json)
# ============================================================


class TestReleasePayloadParsing:
    def test_parse_picks_localization_asset(self):
        payload = {
            "tag_name": "3.0.3",
            "published_at": "2026-06-03T05:48:59Z",
            "html_url": "https://github.com/pinisok/gaku-patcher/releases/tag/3.0.3",
            "assets": [
                {"name": "GakumasLocalify_v3.2.0k.apk",
                 "browser_download_url": "https://example/apk"},
                {"name": "localization.json",
                 "browser_download_url": "https://example/loc.json"},
            ],
        }
        info = _parse_release_payload(payload)
        assert info is not None
        assert info.tag == "3.0.3"
        assert info.asset_url == "https://example/loc.json"

    def test_parse_returns_none_when_asset_missing(self):
        payload = {"tag_name": "3.0.3", "assets": [
            {"name": "apk-only.apk", "browser_download_url": "x"},
        ]}
        assert _parse_release_payload(payload) is None

    def test_parse_returns_none_when_tag_missing(self):
        assert _parse_release_payload({"assets": []}) is None


class TestReleaseTagCache:
    def test_roundtrip(self, tmp_path):
        cache = str(tmp_path / "tag.txt")
        assert load_last_release_tag(cache) is None
        save_release_tag("3.0.3", cache)
        assert load_last_release_tag(cache) == "3.0.3"

    def test_save_trims_whitespace(self, tmp_path):
        cache = str(tmp_path / "tag.txt")
        save_release_tag("  3.0.4  ", cache)
        assert load_last_release_tag(cache) == "3.0.4"


def _release_xlsx(path, rows):
    """Build a localization xlsx whose JP/KR/ID columns are discovered by header
    name. The pandas writer here happens to order cols as (JP, ID, KR); the
    production sheet uses (JP, KR, ID). Both work because the helper resolves
    columns by header value."""
    create_localization_xlsx(path, [
        {0: jp, "ID": key, "번역": kr} for jp, kr, key in rows
    ])


def _read_rows_by_id(xlsx_path):
    """Return {id: (jp, kr)} from an xlsx, looking up cols by header name."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        headers = [c.value for c in ws[1]]
        jp_col = headers.index(0) if 0 in headers else headers.index("0")
        kr_col = headers.index("번역")
        id_col = headers.index("ID")
        rows = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if id_col >= len(row):
                continue
            key = row[id_col]
            if not key:
                continue
            rows[key] = (row[jp_col], row[kr_col])
        return rows
    finally:
        wb.close()


class TestDiffReleaseAgainstXlsx:
    def test_added_changed_removed(self, tmp_path):
        xlsx = str(tmp_path / "loc.xlsx")
        _release_xlsx(xlsx, [
            ("確認", "확인", "ui.ok"),         # unchanged
            ("古い文言", "옛 문구", "ui.old"),  # JP will change
            ("削除対象", "삭제 대상", "ui.removed"),  # not in release
        ])
        release_json = {
            "ui.ok": "確認",
            "ui.old": "新しい文言",
            "ui.new": "新規追加",
        }

        diff = diff_release_against_xlsx(release_json, xlsx)

        assert diff.added == {"ui.new": "新規追加"}
        assert diff.changed_jp == {"ui.old": ("古い文言", "新しい文言")}
        assert diff.removed == ["ui.removed"]
        assert not diff.empty
        assert diff.total == 3

    def test_empty_when_aligned(self, tmp_path):
        xlsx = str(tmp_path / "loc.xlsx")
        _release_xlsx(xlsx, [("確認", "확인", "ui.ok")])
        diff = diff_release_against_xlsx({"ui.ok": "確認"}, xlsx)
        assert diff.empty

    def test_treats_missing_xlsx_as_full_add(self, tmp_path):
        missing = str(tmp_path / "absent.xlsx")
        diff = diff_release_against_xlsx({"a": "あ", "b": "い"}, missing)
        assert diff.added == {"a": "あ", "b": "い"}
        assert not diff.changed_jp and not diff.removed


class TestApplyDiffToXlsx:
    def test_append_change_obsolete(self, tmp_path):
        xlsx = str(tmp_path / "loc.xlsx")
        _release_xlsx(xlsx, [
            ("確認", "확인", "ui.ok"),
            ("古い", "옛", "ui.changed"),
            ("削除", "삭제", "ui.removed"),
        ])
        release_json = {
            "ui.ok": "確認",
            "ui.changed": "新しい",
            "ui.new": "新規",
        }
        diff = diff_release_against_xlsx(release_json, xlsx)
        apply_diff_to_xlsx(release_json, diff, xlsx)

        rows_by_id = _read_rows_by_id(xlsx)

        assert rows_by_id["ui.ok"] == ("確認", "확인")
        assert rows_by_id["ui.changed"][0] == "新しい"
        assert rows_by_id["ui.removed"][0].startswith(OBSOLETE_MARKER)
        assert rows_by_id["ui.removed"][1] == "삭제"  # KR preserved
        assert rows_by_id["ui.new"][0] == "新規"

    def test_obsolete_marker_idempotent(self, tmp_path):
        xlsx = str(tmp_path / "loc.xlsx")
        _release_xlsx(xlsx, [("削除", "삭제", "ui.removed")])

        release_json: dict = {}
        diff1 = diff_release_against_xlsx(release_json, xlsx)
        apply_diff_to_xlsx(release_json, diff1, xlsx)
        # second apply should not double-prefix
        diff2 = diff_release_against_xlsx(release_json, xlsx)
        apply_diff_to_xlsx(release_json, diff2, xlsx)

        rows = _read_rows_by_id(xlsx)
        assert rows["ui.removed"][0].count(OBSOLETE_MARKER) == 1


class TestAppendReleaseNotes:
    def test_creates_file_and_section(self, tmp_path):
        notes = str(tmp_path / "RELEASE_NOTES.md")
        release = ReleaseInfo(
            tag="3.0.3",
            published_at="2026-06-03T05:48:59Z",
            asset_url="https://example/loc.json",
            html_url="https://github.com/pinisok/gaku-patcher/releases/tag/3.0.3",
        )
        diff = LocalizationDiff(
            added={"ui.new": "新規"},
            changed_jp={"ui.changed": ("旧", "新")},
            removed=["ui.removed"],
        )
        append_release_notes(release, diff, notes)

        with open(notes, "r", encoding="utf-8") as f:
            text = f.read()
        assert text.startswith("# Localization Release Notes")
        assert "## 3.0.3" in text
        assert "Source: https://github.com/pinisok/gaku-patcher/releases/tag/3.0.3" in text
        assert "+1 added / ~1 jp-changed / -1 removed" in text
        assert "`ui.new`" in text
        assert "`ui.changed`" in text
        assert "`ui.removed`" in text

    def test_skips_when_tag_already_present(self, tmp_path):
        notes = str(tmp_path / "RELEASE_NOTES.md")
        release = ReleaseInfo(tag="3.0.3", published_at="", asset_url="", html_url="")
        diff = LocalizationDiff(added={"a": "あ"})
        append_release_notes(release, diff, notes)
        size_after_first = os.path.getsize(notes)

        append_release_notes(release, diff, notes)
        assert os.path.getsize(notes) == size_after_first

    def test_prepends_new_release_above_old(self, tmp_path):
        notes = str(tmp_path / "RELEASE_NOTES.md")
        old = ReleaseInfo(tag="3.0.2", published_at="", asset_url="", html_url="")
        new = ReleaseInfo(tag="3.0.3", published_at="", asset_url="", html_url="")
        append_release_notes(old, LocalizationDiff(added={"a": "あ"}), notes)
        append_release_notes(new, LocalizationDiff(added={"b": "い"}), notes)

        with open(notes, "r", encoding="utf-8") as f:
            text = f.read()
        # 3.0.3 must appear before 3.0.2 in the file
        assert text.index("## 3.0.3") < text.index("## 3.0.2")


class TestUpdateOriginalToDriveOrchestration:
    """End-to-end orchestration with all external calls mocked."""

    def test_gates_on_release_tag(self, tmp_path, monkeypatch):
        # Set up cache that already saw 3.0.3.
        cache = tmp_path / "tag.txt"
        cache.write_text("3.0.3", encoding="utf-8")

        monkeypatch.setattr(
            "scripts.localization.localization_release.fetch_latest_release",
            lambda *a, **k: ReleaseInfo(tag="3.0.3", published_at="", asset_url="x", html_url=""),
        )
        monkeypatch.setattr(
            "scripts.localization.localization_release.load_last_release_tag",
            lambda *a, **k: "3.0.3",
        )
        called = {"download": False}
        def _no_download(*a, **k):
            called["download"] = True
            return {}
        monkeypatch.setattr(
            "scripts.localization.localization_release.download_release_json",
            _no_download,
        )

        file_list, warnings = UpdateOriginalToDrive()
        assert file_list == []
        assert warnings == {}
        assert called["download"] is False

    def test_applies_diff_and_writes_notes(self, tmp_path, monkeypatch):
        # Existing drive xlsx.
        xlsx = tmp_path / "localization.xlsx"
        _release_xlsx(str(xlsx), [("旧", "옛", "ui.k")])

        notes = tmp_path / "RELEASE_NOTES.md"
        cache = tmp_path / "tag.txt"

        monkeypatch.setattr(
            "scripts.localization.LOCALIZATION_DRIVE_PATH", str(xlsx)
        )
        monkeypatch.setattr(
            "scripts.localization_release.LOCALIZATION_RELEASE_NOTES_PATH", str(notes)
        )
        monkeypatch.setattr(
            "scripts.localization_release.LOCALIZATION_RELEASE_CACHE_FILE", str(cache)
        )

        release = ReleaseInfo(tag="3.0.3", published_at="2026-06-03T05:48:59Z",
                              asset_url="https://example/loc.json", html_url="https://example/r")
        monkeypatch.setattr(
            "scripts.localization.localization_release.fetch_latest_release",
            lambda *a, **k: release,
        )
        monkeypatch.setattr(
            "scripts.localization.localization_release.download_release_json",
            lambda *a, **k: {"ui.k": "新", "ui.added": "追加"},
        )

        file_list, warnings = UpdateOriginalToDrive()

        assert len(file_list) == 1
        assert "localization.xlsx" in file_list[0][2]
        assert warnings, "should surface warnings for the sheet log"
        assert cache.read_text(encoding="utf-8") == "3.0.3"
        assert notes.exists()
        assert "## 3.0.3" in notes.read_text(encoding="utf-8")

        # Verify xlsx was actually mutated.
        rows = _read_rows_by_id(str(xlsx))
        assert rows["ui.k"][0] == "新"
        assert rows["ui.added"][0] == "追加"

