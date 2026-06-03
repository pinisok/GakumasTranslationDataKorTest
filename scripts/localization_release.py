"""
Localization release sync.

The JP localization source-of-truth lives in the latest release of
`pinisok/gaku-patcher` as a `localization.json` asset (gated by release tag).

This module:
  1. Polls the GitHub release endpoint.
  2. Gates work on the cached tag — skip when the latest tag was already seen.
  3. Downloads the asset to a temp path.
  4. Diffs JSON (id → JP text) vs. the existing drive xlsx (cols: JP, KR, ID).
  5. Mutates the xlsx in place (openpyxl, preserves formatting) — appends new
     keys, updates JP for changed keys, marks removed keys as obsolete.
  6. Appends a markdown section to RELEASE_NOTES.md.

UpdateOriginalToDrive in localization.py orchestrates the pipeline.
"""

from __future__ import annotations

import json
import os
import urllib.request
import urllib.error
from dataclasses import dataclass, field
from datetime import datetime, timezone
from typing import Iterable

import openpyxl

from .log import LOG_DEBUG, LOG_INFO, LOG_WARN, LOG_ERROR
from .paths import (
    LOCALIZATION_RELEASE_API_URL,
    LOCALIZATION_RELEASE_ASSET_NAME,
    LOCALIZATION_RELEASE_CACHE_FILE,
    LOCALIZATION_SOURCE_JSON_PATH,
    LOCALIZATION_RELEASE_NOTES_PATH,
    LOCALIZATION_DRIVE_PATH,
)


JP_COL_HEADER = 0          # openpyxl reads the unnamed first header as int 0
KR_COL_HEADER = "번역"
ID_COL_HEADER = "ID"
OBSOLETE_MARKER = "[OBSOLETE]"

HTTP_TIMEOUT_SECONDS = 30
USER_AGENT = "GakuToolkit-localization-sync/1.0"


# ============================================================
# Data classes
# ============================================================


@dataclass(frozen=True)
class ReleaseInfo:
    tag: str
    published_at: str
    asset_url: str
    html_url: str


@dataclass
class LocalizationDiff:
    """Result of comparing release JSON against the drive xlsx."""

    added: dict[str, str] = field(default_factory=dict)        # id → JP
    changed_jp: dict[str, tuple[str, str]] = field(default_factory=dict)  # id → (old_jp, new_jp)
    removed: list[str] = field(default_factory=list)            # ids only in xlsx

    @property
    def empty(self) -> bool:
        return not (self.added or self.changed_jp or self.removed)

    @property
    def total(self) -> int:
        return len(self.added) + len(self.changed_jp) + len(self.removed)


# ============================================================
# Cache (last seen release tag)
# ============================================================


def load_last_release_tag(cache_file: str | None = None) -> str | None:
    # Resolve module-level constant at call time so tests can monkeypatch it.
    target = cache_file if cache_file is not None else LOCALIZATION_RELEASE_CACHE_FILE
    if not os.path.exists(target):
        return None
    try:
        with open(target, "r", encoding="utf-8") as f:
            tag = f.read().strip()
        return tag or None
    except OSError as e:
        LOG_WARN(2, f"Failed to read localization release cache: {e}")
        return None


def save_release_tag(tag: str, cache_file: str | None = None) -> None:
    target = cache_file if cache_file is not None else LOCALIZATION_RELEASE_CACHE_FILE
    os.makedirs(os.path.dirname(target) or ".", exist_ok=True)
    with open(target, "w", encoding="utf-8") as f:
        f.write(tag.strip())


# ============================================================
# GitHub release fetch
# ============================================================


def fetch_latest_release(api_url: str | None = None) -> ReleaseInfo | None:
    """Hit the releases/latest endpoint. Returns None on network / API failure
    so the caller can skip rather than crash the pipeline."""
    api_url = api_url if api_url is not None else LOCALIZATION_RELEASE_API_URL
    req = urllib.request.Request(api_url, headers={
        "Accept": "application/vnd.github+json",
        "User-Agent": USER_AGENT,
    })
    token = os.environ.get("GITHUB_TOKEN") or os.environ.get("GH_TOKEN")
    if token:
        req.add_header("Authorization", f"Bearer {token}")

    try:
        with urllib.request.urlopen(req, timeout=HTTP_TIMEOUT_SECONDS) as resp:
            payload = json.loads(resp.read().decode("utf-8"))
    except (urllib.error.URLError, TimeoutError, ValueError) as e:
        LOG_ERROR(2, f"Failed to fetch latest release from {api_url}: {e}")
        return None

    return _parse_release_payload(payload)


def _parse_release_payload(payload: dict) -> ReleaseInfo | None:
    tag = payload.get("tag_name") or payload.get("name")
    if not tag:
        LOG_ERROR(2, "Release payload missing tag_name")
        return None

    asset_url = ""
    for asset in payload.get("assets", []):
        if asset.get("name") == LOCALIZATION_RELEASE_ASSET_NAME:
            asset_url = asset.get("browser_download_url", "")
            break
    if not asset_url:
        LOG_ERROR(2, f"Release {tag} has no {LOCALIZATION_RELEASE_ASSET_NAME} asset")
        return None

    return ReleaseInfo(
        tag=str(tag),
        published_at=str(payload.get("published_at") or ""),
        asset_url=asset_url,
        html_url=str(payload.get("html_url") or ""),
    )


def download_release_json(asset_url: str, dest_path: str | None = None) -> dict[str, str] | None:
    """Download the localization.json asset to disk and parse it.

    Returns id→JP dict on success; None on failure."""
    dest_path = dest_path if dest_path is not None else LOCALIZATION_SOURCE_JSON_PATH
    os.makedirs(os.path.dirname(dest_path) or ".", exist_ok=True)
    req = urllib.request.Request(asset_url, headers={"User-Agent": USER_AGENT})
    try:
        with urllib.request.urlopen(req, timeout=HTTP_TIMEOUT_SECONDS) as resp:
            raw = resp.read()
    except (urllib.error.URLError, TimeoutError) as e:
        LOG_ERROR(2, f"Failed to download {asset_url}: {e}")
        return None

    try:
        data = json.loads(raw.decode("utf-8"))
    except (ValueError, UnicodeDecodeError) as e:
        LOG_ERROR(2, f"Release asset is not valid JSON: {e}")
        return None

    if not isinstance(data, dict):
        LOG_ERROR(2, f"Release asset JSON must be an object, got {type(data).__name__}")
        return None

    # Cache the raw file to disk for inspection / fallback.
    with open(dest_path, "wb") as f:
        f.write(raw)

    # Normalize: keep only string values; ignore other types defensively.
    return {str(k): str(v) for k, v in data.items() if isinstance(v, str)}


# ============================================================
# Diff & xlsx mutation
# ============================================================


def _resolve_columns(headers: list) -> tuple[int, int, int]:
    """Locate (jp_col, kr_col, id_col) zero-based indexes. Raises ValueError if missing."""
    jp_col = kr_col = id_col = -1
    for i, h in enumerate(headers):
        if id_col == -1 and h == ID_COL_HEADER:
            id_col = i
        elif kr_col == -1 and h == KR_COL_HEADER:
            kr_col = i
        elif jp_col == -1:
            # First non-ID, non-KR header is the JP source column.
            # openpyxl returns the unnamed first cell as integer 0 or "0".
            if h == JP_COL_HEADER or h == "0" or i == 0:
                jp_col = i
    if id_col == -1 or kr_col == -1 or jp_col == -1:
        raise ValueError(f"Localization xlsx is missing required columns. Headers: {headers}")
    return jp_col, kr_col, id_col


def diff_release_against_xlsx(release_data: dict[str, str], xlsx_path: str) -> LocalizationDiff:
    """Compute additions / JP-changes / removals between release JSON and existing xlsx."""
    diff = LocalizationDiff()

    if not os.path.exists(xlsx_path):
        diff.added = dict(release_data)
        return diff

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        headers = [c.value for c in ws[1]]
        jp_col, _, id_col = _resolve_columns(headers)

        existing: dict[str, str] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if id_col >= len(row) or jp_col >= len(row):
                continue
            key = row[id_col]
            if not isinstance(key, str) or not key:
                continue
            jp = row[jp_col] if isinstance(row[jp_col], str) else ""
            existing[key] = jp
    finally:
        wb.close()

    for key, jp in release_data.items():
        if key not in existing:
            diff.added[key] = jp
        elif existing[key] != jp:
            diff.changed_jp[key] = (existing[key], jp)

    for key in existing:
        if key not in release_data:
            diff.removed.append(key)
    diff.removed.sort()
    return diff


def apply_diff_to_xlsx(release_data: dict[str, str], diff: LocalizationDiff, xlsx_path: str) -> None:
    """Mutate xlsx in place — append new rows, refresh JP for changed rows, mark removed.

    Preserves all existing formatting (uses openpyxl, never pandas/xlsxwriter)."""
    if diff.empty:
        LOG_DEBUG(2, "No localization diff to apply")
        return

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    jp_col, kr_col, id_col = _resolve_columns(headers)

    row_index_by_id: dict[str, int] = {}
    for row_idx in range(2, ws.max_row + 1):
        key_cell = ws.cell(row=row_idx, column=id_col + 1)
        if isinstance(key_cell.value, str) and key_cell.value:
            row_index_by_id[key_cell.value] = row_idx

    # Refresh JP for changed entries.
    for key, (_old, new_jp) in diff.changed_jp.items():
        row_idx = row_index_by_id.get(key)
        if row_idx is None:
            continue
        ws.cell(row=row_idx, column=jp_col + 1, value=new_jp)

    # Mark removed entries — prepend marker into JP if not already marked.
    for key in diff.removed:
        row_idx = row_index_by_id.get(key)
        if row_idx is None:
            continue
        cell = ws.cell(row=row_idx, column=jp_col + 1)
        existing_val = cell.value if isinstance(cell.value, str) else ""
        if not existing_val.startswith(OBSOLETE_MARKER):
            cell.value = f"{OBSOLETE_MARKER} {existing_val}".strip()

    # Append new keys.
    next_row = ws.max_row + 1 if ws.max_row else 2
    for key, jp in diff.added.items():
        ws.cell(row=next_row, column=jp_col + 1, value=jp)
        ws.cell(row=next_row, column=kr_col + 1, value="")
        ws.cell(row=next_row, column=id_col + 1, value=key)
        next_row += 1

    wb.save(xlsx_path)
    LOG_INFO(2, f"localization.xlsx updated — added={len(diff.added)} "
                f"changed_jp={len(diff.changed_jp)} removed={len(diff.removed)}")


# ============================================================
# Release notes
# ============================================================


def append_release_notes(release: ReleaseInfo, diff: LocalizationDiff,
                         notes_path: str | None = None,
                         sample_limit: int = 20) -> None:
    """Append a markdown section for the release. Idempotent: skips if the tag
    section already exists at the top."""
    notes_path = notes_path if notes_path is not None else LOCALIZATION_RELEASE_NOTES_PATH
    os.makedirs(os.path.dirname(notes_path) or ".", exist_ok=True)

    section = _format_release_section(release, diff, sample_limit)

    existing = ""
    if os.path.exists(notes_path):
        with open(notes_path, "r", encoding="utf-8") as f:
            existing = f.read()
        if f"## {release.tag}" in existing:
            LOG_DEBUG(2, f"Release notes already mention {release.tag} — skipping")
            return

    header = "# Localization Release Notes\n\n"
    if existing.startswith(header):
        body = existing[len(header):]
    else:
        body = existing
    with open(notes_path, "w", encoding="utf-8") as f:
        f.write(header + section + ("\n" + body if body else ""))


def _format_release_section(release: ReleaseInfo, diff: LocalizationDiff, sample_limit: int) -> str:
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    lines: list[str] = []
    lines.append(f"## {release.tag} ({today})")
    if release.html_url:
        lines.append(f"- Source: {release.html_url}")
    if release.published_at:
        lines.append(f"- Published: {release.published_at}")
    lines.append(
        f"- Changes: +{len(diff.added)} added / ~{len(diff.changed_jp)} jp-changed / "
        f"-{len(diff.removed)} removed"
    )

    if diff.added:
        lines.append("")
        lines.append("### Added keys")
        for key in list(diff.added.keys())[:sample_limit]:
            lines.append(f"- `{key}`")
        if len(diff.added) > sample_limit:
            lines.append(f"- … +{len(diff.added) - sample_limit} more")

    if diff.changed_jp:
        lines.append("")
        lines.append("### JP changed")
        for key in list(diff.changed_jp.keys())[:sample_limit]:
            lines.append(f"- `{key}`")
        if len(diff.changed_jp) > sample_limit:
            lines.append(f"- … +{len(diff.changed_jp) - sample_limit} more")

    if diff.removed:
        lines.append("")
        lines.append("### Removed keys")
        for key in diff.removed[:sample_limit]:
            lines.append(f"- `{key}`")
        if len(diff.removed) > sample_limit:
            lines.append(f"- … +{len(diff.removed) - sample_limit} more")

    lines.append("")
    return "\n".join(lines)


# ============================================================
# Summary helpers (for gspread / log)
# ============================================================


def summarize_diff(release: ReleaseInfo, diff: LocalizationDiff) -> str:
    return (
        f"localization release {release.tag}: "
        f"+{len(diff.added)} / ~{len(diff.changed_jp)} / -{len(diff.removed)}"
    )
