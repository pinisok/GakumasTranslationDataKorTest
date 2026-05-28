"""MasterDB2 translation helper: find untranslated items and apply translations.

Usage:
    # Find untranslated items
    python -m scripts.masterdb2_translate scan

    # Find untranslated items and export to JSON
    python -m scripts.masterdb2_translate scan --export /tmp/untranslated.json

    # Apply translations from JSON file
    python -m scripts.masterdb2_translate apply translations.json

    # Show character tone samples for translation reference
    python -m scripts.masterdb2_translate tone [char_code]

    # Check produceDescription concat naturalness (JP-faithful vs KR-introduced)
    python -m scripts.masterdb2_translate concat-check
    python -m scripts.masterdb2_translate concat-check --file ProduceCard

JSON format for apply:
    {"FileName": {"原文テキスト": "번역 텍스트", ...}, ...}
"""

import argparse
import json
import os
import re
import sys

import openpyxl
import pandas as pd

from . import paths as _paths


def _read_xlsx(file_path: str) -> pd.DataFrame:
    """Read an xlsx file with the same settings used by the pipeline."""
    df = pd.read_excel(
        file_path,
        na_values="ERROR_NA_VALUE",
        keep_default_na=False,
        na_filter=False,
        engine="openpyxl",
    )
    df.fillna("ERROR_NA_VALUE", inplace=True)
    return df


def scan_untranslated(export_path: str | None = None) -> dict[str, list[dict]]:
    """Scan all masterDB2 xlsx files for untranslated entries.

    Returns dict: {filename: [{keys, id, text}, ...]}
    """
    drive_path = _paths.MASTERDB2_DRIVE_PATH
    result: dict[str, list[dict]] = {}
    total = 0

    for f in sorted(os.listdir(drive_path)):
        if not f.endswith(".xlsx"):
            continue
        name = f[:-5]
        try:
            df = _read_xlsx(os.path.join(drive_path, f))
            if "원문" not in df.columns or "번역" not in df.columns:
                continue
            mask = (df["원문"].astype(str).str.strip() != "") & (
                df["번역"].astype(str).str.strip() == ""
            )
            rows = df[mask]
            if len(rows) == 0:
                continue

            items = []
            for _, row in rows.iterrows():
                keys = "|".join(
                    str(row[c]) for c in df.columns if c.startswith("KEY VALUE")
                )
                items.append(
                    {
                        "keys": keys,
                        "id": str(row.get("ID", "")),
                        "text": str(row["원문"]),
                    }
                )
            result[name] = items
            total += len(items)
            print(f"  {name}: {len(items)}개")
        except Exception as e:
            print(f"  ERROR {name}: {e}", file=sys.stderr)

    print(f"\n총 미번역: {total}개 ({len(result)}개 파일)")

    if export_path:
        with open(export_path, "w", encoding="utf-8") as fp:
            json.dump(result, fp, ensure_ascii=False, indent=2)
        print(f"Exported to {export_path}")

    return result


def apply_translations(trans_path: str) -> None:
    """Apply translations from a JSON file to xlsx files, preserving formatting.

    JSON format: {"FileName": {"原文": "번역", ...}, ...}
    Uses openpyxl to modify cells directly, keeping all formatting intact.
    """
    with open(trans_path, "r", encoding="utf-8") as fp:
        data: dict[str, dict[str, str]] = json.load(fp)

    drive_path = _paths.MASTERDB2_DRIVE_PATH
    total_applied = 0

    for file_name, trans_map in data.items():
        xlsx_path = os.path.join(drive_path, file_name + ".xlsx")
        if not os.path.exists(xlsx_path):
            print(f"  SKIP: {file_name}.xlsx not found")
            continue

        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active

        header_row = [cell.value for cell in ws[1]]
        try:
            orig_col = header_row.index("원문") + 1
            trans_col = header_row.index("번역") + 1
        except ValueError:
            print(f"  SKIP: {file_name} - missing 원문/번역 columns")
            continue

        file_count = 0
        for row in range(2, ws.max_row + 1):
            orig_cell = ws.cell(row=row, column=orig_col)
            trans_cell = ws.cell(row=row, column=trans_col)

            orig_val = str(orig_cell.value) if orig_cell.value is not None else ""
            trans_val = (
                str(trans_cell.value).strip() if trans_cell.value is not None else ""
            )

            if trans_val == "" and orig_val in trans_map:
                trans_cell.value = trans_map[orig_val]
                file_count += 1

        if file_count > 0:
            wb.save(xlsx_path)
            total_applied += file_count
            print(f"  OK: {file_name} - {file_count}개 적용")
        else:
            print(f"  SKIP: {file_name} - no matches")

    print(f"\n총 {total_applied}개 번역 적용")


def show_tone(char_code: str | None = None) -> None:
    """Show existing translated messages per character for tone reference.

    If char_code is given, show only that character. Otherwise show all.
    """
    drive_path = _paths.MASTERDB2_DRIVE_PATH
    xlsx_path = os.path.join(drive_path, "CharacterPushMessage.xlsx")
    if not os.path.exists(xlsx_path):
        print("CharacterPushMessage.xlsx not found")
        return

    df = _read_xlsx(xlsx_path)
    if "KEY VALUE 0" not in df.columns:
        print("Unexpected column structure")
        return

    chars = (
        [char_code]
        if char_code
        else sorted(df["KEY VALUE 0"].astype(str).unique())
    )

    for char in chars:
        if len(char) != 4 or not char.isalpha():
            continue
        mask = (
            (df["KEY VALUE 0"].astype(str) == char)
            & (df["ID"] == "message")
            & (df["번역"].astype(str).str.strip() != "")
        )
        rows = df[mask]
        if len(rows) == 0:
            continue
        print(f"\n=== {char} ({len(rows)}개) ===")
        for _, row in rows.head(6).iterrows():
            typ = str(row.get("KEY VALUE 1", ""))
            orig = str(row["원문"])[:80]
            trans = str(row["번역"])[:80]
            print(f"  [{typ}]")
            print(f"    JP: {orig}")
            print(f"    KR: {trans}")


# ============================================================
# Fragment Concat 자연스러움 검증
# ============================================================
#
# MasterDB2의 produceDescription은 짧은 fragment들이 record 별로 concat되어 UI에
# 표시된다. 일본어는 한자가 시각적으로 붙어도 자연스럽지만, 한국어는 명시적
# 공백/조사 처리가 필요해서 fragment 경계에서 회귀가 자주 발생한다.
#
# 검증 패턴 (모두 produceDescriptions concat 결과 기준):
#   - doublespace: 더블 스페이스. JP source가 ' '+' ' 페어인 경우 JP-faithful로
#                  분류 (유지 가능), 그 외는 KR-introduced (수정 필요).
#   - noun_compress: "원기증가" 같이 상태명 + 동작 명사가 붙음. 띄어쓰기 필요.
#   - 호조_의: "호조 의 X%" 같이 의(possessive)가 fragment 단독으로 분리되어
#              어색하게 띄어진 경우.
#   - particle_compress: "호조이/강기가" 같이 받침에 안 맞는 particle이
#                        fragment 시작에 있어 잘못 concat된 경우.
#   - card_no_space_사용: 카드명 + 사용 띄어쓰기 누락 (에이에이오!사용 후).
#   - x_됐을때: 상태명 + 됐을 때 띄어쓰기 누락 (전력됐을 때).
#   - comma_no_space: "일 때,X" 같이 콤마 뒤 공백 누락.
#   - leading_no_space_있는N: "있는자연스러운" 같이 "있는" + 명사 띄어쓰기 누락.
#   - 경감_잔존: "소비 체력 경감" (감소로 통일해야 함).
#   - 개시시_잔존: "개시시" (시작 시로 통일해야 함).
#   - 버림_패_잔존: "버림 패" (버린 패로 통일해야 함).
#   - 인_경우_잔존: "~인 경우," (~일 때, 로 통일해야 함).

_CONCAT_CHECK_PATTERNS = {
    "doublespace": re.compile(r"(?<![<])  +"),
    "noun_compress": re.compile(
        r"(원기|집중|호인상|호조|의욕|강기|절호조|매력|전력|온존)(증가|상승|감소)량?[^:가-힣\)\(]"
    ),
    "호조_의": re.compile(r"(호조|집중|의욕|호인상|강기|절호조|원기|전력|매력) 의"),
    # 받침 X (vowel-ending) status names followed by particle 이 — should be 가 instead
    "particle_compress": re.compile(r"(?<![가-힣])(호조|절호조|강기|원기)이[\s<,]"),
    "card_no_space_사용": re.compile(
        r"(에이에이오!|럭키♪|하나ー둘!?|해피ー♪|한숨|페이스 조절|슈프레히콜|스타트 대시|스탠드 플레이|스포트라이트|데이드리밍|반짝임|연출 계획|존재감|소원의 힘|정신 통일|모티베이션|성취감|절차탁마|프라이드|첫인상|액티브 스킬 카드)(?<!\s)사용"
    ),
    "x_됐을때": re.compile(r"(?<=[가-힣])됐을 때"),
    "comma_no_space": re.compile(r"[일때경우],[가-힣]"),
    "leading_no_space_있는N": re.compile(r"있는[가-힣]"),
    "경감_잔존": re.compile(r"소비 체력 경감|체력 감소량 경감"),
    "개시시_잔존": re.compile(r"개시시"),
    "버림_패_잔존": re.compile(r"버림 패"),
    "인_경우_잔존": re.compile(r"인 경우,"),
}


def _collect_descs_concat(record: dict) -> str:
    """Concat all *Descriptions field texts in a record (mimics UI rendering)."""
    chunks = []
    for key in (
        "produceDescriptions",
        "customizeProduceDescriptions",
        "playEffectProduceDescriptions",
        "playProduceDescriptions",
    ):
        for d in record.get(key, []) or []:
            t = d.get("text", "") if isinstance(d, dict) else ""
            if isinstance(t, str):
                chunks.append(t)
    return "".join(chunks)


def concat_check(target_file: str | None = None) -> int:
    """Scan output JSON for fragment concat naturalness issues.

    Classifies doublespace into JP-faithful (preserved from JP source) vs
    KR-introduced (translation bug). Returns total KR-introduced issue count.
    """
    output_json_dir = "output/local-files/masterTrans"
    jp_source_dir = "res/masterdb/gakumasu-diff/json"

    if not os.path.isdir(output_json_dir):
        print(f"❌ output JSON not found: {output_json_dir}", file=sys.stderr)
        return -1

    total_kr_introduced = 0
    total_jp_faithful = 0
    per_file: dict[str, dict[str, int]] = {}

    for fn in sorted(os.listdir(output_json_dir)):
        if not fn.endswith(".json"):
            continue
        fname = fn[:-5]
        if target_file and fname != target_file:
            continue
        fp = os.path.join(output_json_dir, fn)
        try:
            with open(fp, encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            continue

        # JP source for doublespace classification
        jp_map: dict[str, dict] = {}
        jp_path = os.path.join(jp_source_dir, fn)
        if os.path.exists(jp_path):
            try:
                with open(jp_path, encoding="utf-8") as f:
                    jp_data = json.load(f)
                jp_map = {
                    r.get("id", ""): r for r in jp_data.get("data", []) if isinstance(r, dict)
                }
            except Exception:
                jp_map = {}

        file_issues: dict[str, int] = {k: 0 for k in _CONCAT_CHECK_PATTERNS}
        for rec in data.get("data", []):
            if not isinstance(rec, dict):
                continue
            rid = rec.get("id", "")
            kr_concat = _collect_descs_concat(rec)
            if not kr_concat:
                continue
            jp_concat = _collect_descs_concat(jp_map.get(rid, {}))

            # For doublespace: classify JP-faithful if either JP concat has " +"
            # OR KR descriptors have adjacent space-only fragments (JP source quirk
            # preserved when JP json wasn't available).
            def _is_doublespace_jp_faithful() -> bool:
                if re.search(r"  +", jp_concat):
                    return True
                # Fallback: KR has adjacent ' '+' ' fragments (JP source structure)
                kr_descs_all = []
                for k2 in (
                    "produceDescriptions",
                    "customizeProduceDescriptions",
                    "playEffectProduceDescriptions",
                    "playProduceDescriptions",
                ):
                    kr_descs_all.extend(rec.get(k2, []) or [])
                for i in range(len(kr_descs_all) - 1):
                    t1 = kr_descs_all[i].get("text", "") if isinstance(kr_descs_all[i], dict) else ""
                    t2 = kr_descs_all[i + 1].get("text", "") if isinstance(kr_descs_all[i + 1], dict) else ""
                    if t1 == " " and t2 == " ":
                        return True
                return False

            for k, pat in _CONCAT_CHECK_PATTERNS.items():
                if not pat.search(kr_concat):
                    continue
                file_issues[k] += 1
                if k == "doublespace" and _is_doublespace_jp_faithful():
                    total_jp_faithful += 1
                else:
                    total_kr_introduced += 1

        if sum(file_issues.values()) > 0:
            per_file[fname] = file_issues

    if not per_file:
        print("✅ 검증 통과: KR-introduced concat 자연스러움 issue 없음.")
        if total_jp_faithful:
            print(f"   (JP-faithful {total_jp_faithful}건은 JP 원본 구조에 충실하므로 보존)")
        return 0

    print("=== Concat 자연스러움 issue 요약 ===")
    for fname, issues in sorted(per_file.items()):
        parts = [f"{k}={v}" for k, v in issues.items() if v]
        print(f"  {fname}: {', '.join(parts)}")
    print(f"\nKR-introduced: {total_kr_introduced} (수정 필요)")
    print(f"JP-faithful: {total_jp_faithful} (JP 원본 구조 보존)")
    return total_kr_introduced


def main() -> None:
    parser = argparse.ArgumentParser(description="MasterDB2 translation helper")
    sub = parser.add_subparsers(dest="command")

    scan_p = sub.add_parser("scan", help="Scan for untranslated items")
    scan_p.add_argument("--export", help="Export to JSON file")

    apply_p = sub.add_parser("apply", help="Apply translations from JSON")
    apply_p.add_argument("file", help="JSON file with translations")

    tone_p = sub.add_parser("tone", help="Show character tone samples")
    tone_p.add_argument("char", nargs="?", help="Character code (e.g. amao)")

    cc_p = sub.add_parser(
        "concat-check",
        help="Verify produceDescription concat naturalness (JP-faithful vs KR-introduced)",
    )
    cc_p.add_argument("--file", help="Restrict to a single JSON file (e.g. ProduceCard)")

    args = parser.parse_args()

    if args.command == "scan":
        scan_untranslated(args.export)
    elif args.command == "apply":
        apply_translations(args.file)
    elif args.command == "tone":
        show_tone(args.char)
    elif args.command == "concat-check":
        code = concat_check(args.file)
        sys.exit(1 if code > 0 else 0)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
